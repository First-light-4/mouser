from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import requests
import time
import openpyxl

class BS4:

    def url(self, url):
        self.url = url

    def get_products(self, name):
        headers = {
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15",
        }

    
        req = requests.get(self.url, headers=headers)
        src = req.text
        with open(f"{name}.html", "w", encoding="utf-8") as file:
            file.write(src)

    def read_html(self, name):

        with open(f"{name}.html", "r", encoding="utf-8") as file:
            self.src = file.read()
        return self.src
    

class Selenium:
    def __init__(self, url):
        self.url = url
        self.data = []
        self.path_webdriver = r'C:\Users\vsavc\PycharmProjects\chrome-win32'


    # подключение к драйверу
    def get_products(self):
        path_to_chrome = fr"{self.path_webdriver}\chrome.exe"
        options = Options()
        options.binary_location = path_to_chrome
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36")
        options.add_argument("disable-blink-features=AutomationControlled")

        service = Service(fr"{self.path_webdriver}\chromedriver.exe") 
        service.start()

        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.get(self.url)
        self.driver.implicitly_wait(10)
        return self.driver
    
    def get_driver(self):
        return self.driver

    def get_links(self, elements):
        
        self.links = [elem.find_element(By.TAG_NAME, "a").get_attribute('href') for elem in elements if elem.find_element(By.TAG_NAME, "a")]
        return self.links

    
    def download_page(self, name):
        # Получение исходного кода страницы
        page_source = self.driver.page_source

        # сохранение кода
        with open(f"{name}.html", "w", encoding="utf-8") as file:
            file.write(page_source)
    
    def click_bottom(self, class_name):
        # Поиск элемента по классу и выполнение клика
        button = self.driver.find_element_by_class_name(class_name)
        button.click()

    def close_connection(self):
        self.driver.quit()

# получение данных с главной страницы
def get_list_types():
    url = 'https://eu.mouser.com/'
    sel = Selenium(url)
    sel.get_products()
    driver  = sel.get_driver()
    time.sleep(10)
    sel.download_page(name='mouser_main')
    elements = driver.find_elements(By.CSS_SELECTOR, ".container.pl-0.py-0 .dropdown-item")
    print(sel.get_links(elements=elements))
    sel.close_connection()

# поиск всех групп данных
def find_href():
    html_content = BS4().read_html(name='mouser_main')
    soup = BeautifulSoup(html_content, 'html.parser')

    container = soup.find('div', class_='container pl-0 py-0')
    dropdown_items = container.find_all('a', class_='dropdown-item')
    return dropdown_items

# добавление новой информации в таблицу
def add_data_to_excel(file_path, data):

    # Загрузка существующего файла
    workbook = openpyxl.load_workbook(file_path)

    # Выбор активного листа
    sheet = workbook.active
    
    # Поиск первой пустой строки
    empty_row = 1
    while sheet.cell(row=empty_row, column=1).value is not None:
        empty_row += 1

    # Добавление данных в строку
    for col, value in enumerate(data, start=1):
        sheet.cell(row=empty_row, column=col, value=value)
    
    # Сохранение изменений
    workbook.save(file_path)

def get_item_info(row):
    Mfr_Part_No = row.find_elements(By.CSS_SELECTOR, '.mfr-part-num.hidden-xs a')
    Mouser_Part_No = row.find_elements(By)
    Mfr = row.find_elements(By)
    Availability = row.find_elements(By)
    pricing_for_1 = row.find_elements(By)
    pricing_for_10 = row.find_elements(By)
    pricing_for_25 = row.find_elements(By)
    pricing_for_100 = row.find_elements(By)
    href = row.find_elements(By)

    return Mfr_Part_No, Mouser_Part_No, Mfr, Availability, pricing_for_1, pricing_for_10, pricing_for_25, pricing_for_100, href
    

if __name__ == '__main__':
    for item in find_href():
        count = 0
        data = [item.text]
        while True:
            count += 1
            try:
                url = f'https://eu.mouser.com/{item["href"]}/?pg={count}'
                
#
                # открытие страницы 
                sel = Selenium(url)
                sel.get_products()
                driver  = sel.get_driver()
                time.sleep(10)
                tbody = driver.find_element(By.TAG_NAME, 'tbody')
#
                rows = tbody.find_elements(By.TAG_NAME, 'tr')
                
                for row in rows:
                    chapter = data
                    Mfr_Part_No = row.find_elements(By.CSS_SELECTOR, '.mfr-part-num.hidden-xs a').text
                    print (chapter, Mfr_Part_No)    
                    break 
                sel.close_connection()
                break
                # запись информации в xlsx
                add_data_to_excel(file_path='mouser.xlsx', data=item)
            except:
                sel.close_connection()
                break
        break


    
